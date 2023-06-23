from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

#Modifica opciones de webdriver en Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
#Importo excepciones y WebDriverWait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import WebDriverException

from selenium.webdriver.support import expected_conditions as EC #Esto es para elementos clicables

#importo time y pongo tiempo
import time

#Libreria para descargar imagenes
import urllib.request

import os

import openpyxl

from openpyxl.drawing.image import Image

from openpyxl.utils import get_column_letter

from openpyxl.styles import Font


#creo una función para iniciar Chrome
def iniciar_chrome():

    ruta=ChromeDriverManager(path="./chromedriver").install()
    options=Options()
    user_agent= "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
    options.add_argument(f"user-agent={user_agent}")#define un user agent personalizado
    #options.add_argument("--window-size=100x100")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")#chromedriver no muestra nada en terminal
    options.add_argument("--no-default-browser-check")
    options.add_argument("--no-first-run")#evita ejecusción de cierre
    options.add_argument("--no-proxy-server")#usar conx directa
    options.add_argument("--disable-link-features=AutomationControlled")#evitar que detecten que somos bot con selenium
    #parámetros a omitir inicio chromedriver
    exp_opt = [
        'ignore-certificate-errors',
        'enable-logging'
    ]
    options.add_experimental_option("excludeSwitches", exp_opt)
    #parámetros de preferencias
    prefs= {
        "profile.default_content_setting_value.notification" : 2,#0=preguntar 1=permitir 2=no permitir
        
        "credentials_enable_service": False#evita que te pida contraseña google al login
    }
    options.add_experimental_option("prefs", prefs)
   

    #inicia chrome con parametros indicados y devuelve driver

    s=Service(ruta)
    #instanciamos webdriver
    driver=webdriver.Chrome(service=s, options=options)
    return driver



#MÉTODO MAIN
if __name__ == '__main__':

    #Inicio la función que inicia chrome
    # Obtener la ruta del directorio actual del programa
    programa_directorio = os.path.dirname(os.path.abspath(__file__))

    # Ruta de la carpeta donde se guardarán las imágenes
    carpeta_destino = os.path.join(programa_directorio, "imagenes")

    # Ruta de la carpeta donde se guardará el archivo de Excel
    carpeta_excel = os.path.join(programa_directorio, "EXCEL")

    # Crear la carpeta si no existe
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)

    ## Crear la carpeta si no existe
    if not os.path.exists(carpeta_excel):
        os.makedirs(carpeta_excel)

    # Me voy a la sección de zapatos de mujer
    url = "https://www.zara.com/es/es/mujer-zapatos-l1251.html?v1=2289820"
    driver = iniciar_chrome()
    driver.get(url)

    opener = urllib.request.build_opener()
    opener.addheaders = [('User-Agent', 'Mozilla/5.0')]
    urllib.request.install_opener(opener)

    # Realizar el scroll en la página
    SCROLL_PAUSE_TIME = 0.1  # Tiempo de espera entre desplazamientos
    scroll_distance = 500  # Distancia de desplazamiento en píxeles
    last_position = 0
    while True:
        # Desplazarse hasta la posición actual + scroll_distance
        driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
        time.sleep(SCROLL_PAUSE_TIME)

        # Obtener la posición actual de desplazamiento
        current_position = driver.execute_script("return window.pageYOffset;")

        if current_position == last_position:
            # Si la posición actual es igual a la posición anterior, se ha alcanzado el final de la página
            break

        last_position = current_position

    # Buscar los elementos de imagen en la página
    imagenes = driver.find_elements(By.CSS_SELECTOR, "img.media-image__image.media__wrapper--media")

    # Crear un nuevo archivo de Excel o cargar uno existente
    excel_nombre = os.path.join(programa_directorio, "datos_productos.xlsx")
    if os.path.exists(excel_nombre):
        workbook = openpyxl.load_workbook(excel_nombre)
    else:
        workbook = openpyxl.Workbook()

    # Crear un nuevo archivo de Excel o cargar uno existente
    excel_nombre = os.path.join(carpeta_excel, "datos_productos.xlsx")
    if os.path.exists(excel_nombre):
        workbook = openpyxl.load_workbook(excel_nombre)
    else:
        workbook = openpyxl.Workbook()
    # Seleccionar la hoja activa
    sheet = workbook.active

    # Encontrar la última fila en la hoja
    last_row = sheet.max_row




    # Agregar encabezados de columna si el archivo está vacío
    if last_row == 1:
        sheet["A1"] = "Nombre"
        sheet["B1"] = "Precio"
        sheet["C1"] = "Imagen"
        sheet["D1"] = "Enlace"


    # Aumentar el tamaño de letra de las secciones "Nombre" y "Precio"
    sheet['A1'].font = Font(size=14, bold=True)  # Tamaño de letra 14 y negrita para la sección "Nombre"
    sheet['B1'].font = Font(size=14, bold=True)  # Tamaño de letra 14 y negrita para la sección "Precio"
    sheet['C1'].font = Font(size=14, bold=True)  # Tamaño de letra 14 y negrita para la sección "Imagen"
    sheet['D1'].font = Font(size=14, bold=True)

    # Obtener el nombre de cada producto y guardarlo en el archivo de Excel
    productos = driver.find_elements(By.CSS_SELECTOR, "a.product-link._item.product-grid-product-info__name.link")
    for i, producto in enumerate(productos):
        nombre_producto = producto.text
        sheet.cell(row=i+2, column=1).value = nombre_producto



    # Obtener el enlace de cada producto
    productos = driver.find_elements(By.CSS_SELECTOR, "a.product-link.product-grid-product__link")
    enlaces = [producto.get_attribute("href") for producto in productos]
    # Insertar los datos en el archivo de Excel
    for i, enlace in enumerate(enlaces):
        sheet.cell(row=i+2, column=4).value = enlace


   # Obtener los elementos que contienen los precios
    precios = driver.find_elements(By.CSS_SELECTOR, "span.product-grid-product-info__2nd-price-amount.price-current--with-background")

    # Insertar los precios en el archivo de Excel
    for i, precio_elemento in enumerate(precios):
        
        precio_texto = precio_elemento.text
        # Quitar el porcentaje del precio
        precio_texto = precio_texto.split("%")[1].strip()
        sheet.cell(row=i+2, column=2).value = precio_texto


    # Descargar y guardar las imágenes en la carpeta destino y el Excel
    for i, imagen in enumerate(imagenes):
        src = imagen.get_attribute("src")
        nombre_archivo = f"imagen{i+1}.jpg"
        ruta_guardado = os.path.join(carpeta_destino, nombre_archivo)
        urllib.request.urlretrieve(src, ruta_guardado)
        print(f"Descargada la imagen {i+1}")

        # Insertar la imagen en la hoja de cálculo
        img = Image(ruta_guardado)
        img.width = 160
        img.height = 200
        sheet.column_dimensions['C'].width = 160
        sheet.row_dimensions[i + 2].height = 200  # Comienza en la fila 2 (C2)
        sheet.add_image(img, f"C{i + 2}")

    # Ajustar el ancho de las columnas en función del contenido
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo de Excel
    
    workbook.save(excel_nombre)

driver.quit()